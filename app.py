import streamlit as st
import sqlite3
import json
from datetime import datetime, date
from openpyxl import load_workbook
import os

# =====================
# DB・データ読み込み
# =====================

EXCEL_PATH = "kantei_db.xlsx"

@st.cache_data
def load_all_data():
    wb = load_workbook(EXCEL_PATH, read_only=True)

    # 五芒星運命の固定数（年月→運命の固定数・干支・αβ）
    ws = wb['五芒星運命の固定数']
    rows = list(ws.iter_rows(values_only=True))
    meimei_db = {}
    base_year = 1924
    for i, r in enumerate(rows[1:], 1):
        year = base_year + (i - 1) // 12
        eto = r[3]
        month = r[4]
        unmei = r[6]
        ab = r[7]
        if isinstance(unmei, int) and eto and month:
            key = f'{year}{month}'
            meimei_db[key] = {'unmei': unmei, 'eto': eto, 'ab': ab}

    # 星（鑑定数→五芒星・天与神）
    ws = wb['星']
    hoshi_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], int):
            hoshi_db[r[0]] = {'hoshi': r[1], 'tenyoshin': r[2]}

    # 王様の金庫（数→称号）
    ws = wb['王様の金庫']
    okura_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], int):
            okura_db[r[0]] = r[1]

    # 花（鑑定数→花）
    ws = wb['花']
    hana_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], int):
            hana_db[r[0]] = r[1]

    # 風景（鑑定数→風景）
    ws = wb['風景']
    fukei_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], int):
            fukei_db[r[0]] = r[1]

    # 文房具（数→文房具）
    ws = wb['文房具']
    bunbogu_db = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], (int, float)) and r[0] is not None:
            bunbogu_db[int(r[0])] = r[1]

    # ジュメリ（ジュメリベース→ジュメリ）
    ws = wb['ジュメリ']
    jumeri_db = {}
    eto_list = ['子','丑','寅','卯','辰','巳','午','未','申','酉','戌','亥']
    hoshi_list_j = ['センリα','センリβ','マリα','マリβ','ノアα','ノアβ','アカリα','アカリβ','キンリュウα','キンリュウβ','カイリα','カイリβ']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        eto_val = r[0]
        hoshi_val = r[1]
        if eto_val and hoshi_val:
            key = f'{eto_val}{hoshi_val}'
            jumeri_db[key] = f'{hoshi_val}ジュメリ'

    return meimei_db, hoshi_db, okura_db, hana_db, fukei_db, bunbogu_db, jumeri_db


def calc_kantei(birthday: date, meimei_db, hoshi_db, okura_db, hana_db, fukei_db, bunbogu_db, jumeri_db):
    """生年月日から全鑑定値を計算する"""
    result = {}

    # 生年月日を8桁に分解
    bd_str = birthday.strftime('%Y%m%d')
    digits = [int(c) for c in bd_str]  # [Y1,Y2,Y3,Y4,M1,M2,D1,D2]

    # 年月キー（YYYYMM）
    year_month_key = f'{birthday.year}{birthday.month:02d}'

    # 五芒星運命の固定数DB検索
    meimei = meimei_db.get(year_month_key)
    if not meimei:
        return None, f'年月 {year_month_key} のデータがDBに見つかりません（対応範囲: 1924年〜2020年）'

    unmei_fixed = meimei['unmei']
    eto = meimei['eto']
    ab = meimei['ab']
    result['干支'] = eto
    # αβ非表示 = ab
    result['運命の固定数'] = unmei_fixed

    # 日（2桁）
    day = birthday.day

    # 五芒星鑑定数 = 運命の固定数 + 日
    gohoshi_num = unmei_fixed + day
    if gohoshi_num > 60:
        gohoshi_num -= 60
    result['五芒星鑑定数'] = gohoshi_num

    # 星・天与神
    hoshi_data = hoshi_db.get(gohoshi_num, {})
    result['天性鑑定'] = hoshi_data.get('hoshi', '?')
    result['天与神'] = hoshi_data.get('tenyoshin', '?')

    # ジュメリベース・ジュメリ
    jumeri_base = f'{eto}{hoshi_data.get("hoshi","")}{ab}'
    result['ジュメリベース'] = jumeri_base
    result['ジュメリ'] = jumeri_db.get(jumeri_base, '単星')

    # 風景・花
    result['風景占い'] = fukei_db.get(gohoshi_num, '?')
    result['花占い'] = hana_db.get(gohoshi_num, '?')

    # 王様の金庫（生年月日8桁の各数字の合計）
    digit_sum = sum(digits)
    digit_sum_str = str(digit_sum)
    d1 = int(digit_sum_str[0]) if len(digit_sum_str) >= 1 else 0
    d2 = int(digit_sum_str[1]) if len(digit_sum_str) >= 2 else 0
    okura_num = d1 + d2
    result['王様の金庫番号'] = okura_num
    result['王様の金庫'] = okura_db.get(okura_num, '?')

    # 文房具（五芒星鑑定数の2桁目）
    gohoshi_str = str(gohoshi_num)
    bunbogu_key = int(gohoshi_str[1]) if len(gohoshi_str) >= 2 else 0
    result['文房具'] = bunbogu_db.get(bunbogu_key, '?')

    return result, None


# =====================
# SQLite DB初期化
# =====================

def init_db():
    conn = sqlite3.connect('kantei_records.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            user_name TEXT,
            birthday TEXT,
            comment TEXT,
            kantei_result TEXT
        )
    ''')
    conn.commit()
    conn.close()

def save_record(user_name, birthday, comment, result):
    conn = sqlite3.connect('kantei_records.db')
    c = conn.cursor()
    c.execute('''
        INSERT INTO records (created_at, user_name, birthday, comment, kantei_result)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        user_name,
        birthday.strftime('%Y-%m-%d'),
        comment,
        json.dumps(result, ensure_ascii=False)
    ))
    conn.commit()
    conn.close()

def load_records():
    conn = sqlite3.connect('kantei_records.db')
    c = conn.cursor()
    c.execute('SELECT * FROM records ORDER BY created_at DESC')
    rows = c.fetchall()
    conn.close()
    return rows


# =====================
# Streamlit UI
# =====================

st.set_page_config(page_title='統計学鑑定システム', page_icon='⭐', layout='wide')

st.title('⭐ 統計学鑑定システム')
st.markdown('---')

# データ読み込み
try:
    meimei_db, hoshi_db, okura_db, hana_db, fukei_db, bunbogu_db, jumeri_db = load_all_data()
    init_db()
except Exception as e:
    st.error(f'データ読み込みエラー: {e}')
    st.stop()

# タブ
tab1, tab2 = st.tabs(['🔮 鑑定する', '📋 鑑定履歴'])

with tab1:
    st.subheader('鑑定情報を入力してください')

    col1, col2 = st.columns(2)
    with col1:
        user_name = st.text_input('お名前', placeholder='例：山田 花子')
    with col2:
        birthday = st.date_input(
            '生年月日',
            min_value=date(1924, 1, 1),
            max_value=date(2020, 12, 31),
            value=date(1980, 1, 1)
        )

    comment = st.text_area('コメント・メモ', placeholder='相談内容、気になること、状況など', height=100)

    if st.button('✨ 鑑定する', use_container_width=True, type='primary'):
        if not user_name:
            st.warning('お名前を入力してください')
        else:
            result, error = calc_kantei(
                birthday, meimei_db, hoshi_db, okura_db,
                hana_db, fukei_db, bunbogu_db, jumeri_db
            )

            if error:
                st.error(error)
            else:
                # 結果表示
                st.markdown('---')
                st.subheader(f'🌟 {user_name} さんの鑑定結果')
                st.caption(f'生年月日: {birthday.strftime("%Y年%m月%d日")}')

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.markdown('#### 🌀 天性鑑定')
                    st.metric('五芒星', result['天性鑑定'])
                    st.metric('五芒星鑑定数', result['五芒星鑑定数'])
                    st.metric('天与神', result['天与神'])
                    # αβ非表示

                with col2:
                    st.markdown('#### 🔮 運命・ジュメリ')
                    st.metric('干支', result['干支'])
                    st.metric('運命の固定数', result['運命の固定数'])
                    st.metric('ジュメリ', result['ジュメリ'])

                with col3:
                    st.markdown('#### 👑 王様の金庫')
                    st.metric('王様の金庫', result['王様の金庫'])
                    st.metric('文房具', result['文房具'])

                st.markdown('#### 🌸 花・風景')
                col4, col5 = st.columns(2)
                with col4:
                    st.metric('花占い', result['花占い'])
                with col5:
                    st.metric('風景占い', result['風景占い'])

                # DB保存
                save_record(user_name, birthday, comment, result)

with tab2:
    st.subheader('📋 鑑定履歴')

    pw = st.text_input('パスワードを入力してください', type='password')
    if pw == 'Lamat':
        records = load_records()
        if not records:
            st.info('まだ鑑定履歴がありません')
        else:
            for rec in records:
                with st.expander(f'📅 {rec[1]}｜{rec[2]}（{rec[3]}）'):
                    st.write(f'**コメント:** {rec[4] or "なし"}')
                    result_data = json.loads(rec[5])
                    cols = st.columns(4)
                    items = list(result_data.items())
                    for i, (k, v) in enumerate(items):
                        with cols[i % 4]:
                            st.metric(k, v)
    elif pw != '':
        st.error('パスワードが違います')
